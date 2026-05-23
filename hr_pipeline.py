"""HR Attrition pipeline — CLI version.

Generates a synthetic HR attrition dataset, cleans it, builds 4 charts, and
exports a polished Excel workbook with a Dashboard sheet, KPIs, raw data,
and an executive summary.

Usage:
    python hr_pipeline.py
"""
from __future__ import annotations

import os
import warnings
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# Output goes to a local ./output folder relative to this script.
PROJECT_ROOT = Path(__file__).resolve().parent
OUT = PROJECT_ROOT / "output"
OUT.mkdir(exist_ok=True)

np.random.seed(42)


# ══════════════════════════════════════════════
# 1. GENERATE IBM-STYLE HR ATTRITION DATASET
# ══════════════════════════════════════════════
print("▶ Step 1/5 — Generating HR Attrition dataset...")

n = 1470
depts = {
    "Sales": 0.34,
    "Research & Development": 0.48,
    "Human Resources": 0.18,
}
job_roles = {
    "Sales": ["Sales Executive", "Sales Representative", "Manager"],
    "Research & Development": [
        "Research Scientist", "Laboratory Technician", "Healthcare Representative",
        "Manufacturing Director", "Research Director",
    ],
    "Human Resources": ["Human Resources", "Manager"],
}
edu_fields = ["Life Sciences", "Medical", "Marketing", "Technical Degree", "Human Resources", "Other"]

rows = []
for i in range(n):
    dept = np.random.choice(list(depts.keys()), p=list(depts.values()))
    role = np.random.choice(job_roles[dept])
    age = max(18, min(60, int(np.random.normal(37, 9))))
    tenure = np.random.randint(0, min(age - 18, 35) + 1)
    salary = max(10000, min(200000, int(np.random.lognormal(9.2, 0.45))))
    dist = np.random.randint(1, 30)
    overtime = np.random.choice(["Yes", "No"], p=[0.28, 0.72])
    satisf = np.random.randint(1, 5)
    env_sat = np.random.randint(1, 5)
    wlb = np.random.randint(1, 5)
    perf = np.random.choice([3, 4], p=[0.84, 0.16])
    edu = np.random.randint(1, 6)
    edu_field = np.random.choice(edu_fields)
    num_cos = np.random.randint(0, 10)
    last_promo = np.random.randint(0, 15)

    p_attr = 0.16
    if overtime == "Yes":
        p_attr += 0.12
    if satisf <= 2:
        p_attr += 0.10
    if wlb <= 2:
        p_attr += 0.08
    if tenure < 2:
        p_attr += 0.10
    if salary < 30000:
        p_attr += 0.08
    if dist > 20:
        p_attr += 0.05
    if dept == "Sales":
        p_attr += 0.04
    if last_promo > 5:
        p_attr += 0.06
    p_attr = min(p_attr, 0.75)

    attrition = np.random.choice(["Yes", "No"], p=[p_attr, 1 - p_attr])

    rows.append({
        "EmployeeID": 1000 + i, "Age": age, "Attrition": attrition,
        "Department": dept, "JobRole": role, "MonthlyIncome": salary,
        "YearsAtCompany": tenure, "DistanceFromHome": dist, "OverTime": overtime,
        "JobSatisfaction": satisf, "EnvironmentSatisfaction": env_sat,
        "WorkLifeBalance": wlb, "PerformanceRating": perf, "Education": edu,
        "EducationField": edu_field, "NumCompaniesWorked": num_cos,
        "YearsSinceLastPromotion": last_promo,
        "Gender": np.random.choice(["Male", "Female"], p=[0.60, 0.40]),
        "MaritalStatus": np.random.choice(["Single", "Married", "Divorced"], p=[0.32, 0.46, 0.22]),
    })

df_raw = pd.DataFrame(rows)
for col in ["MonthlyIncome", "JobSatisfaction", "YearsAtCompany"]:
    mask = np.random.random(len(df_raw)) < 0.03
    df_raw.loc[mask, col] = np.nan
dupes = df_raw.sample(30, random_state=5)
df_raw = pd.concat([df_raw, dupes], ignore_index=True)
df_raw.to_csv(OUT / "dirty_hr_attrition.csv", index=False)
print(f"   Raw: {len(df_raw)} rows | Missing: {df_raw.isnull().sum().sum()} | Dupes: 30")

# ══════════════════════════════════════════════
# 2. CLEAN
# ══════════════════════════════════════════════
print("▶ Step 2/5 — Cleaning data...")
df = df_raw.drop_duplicates().copy()
df["MonthlyIncome"] = df["MonthlyIncome"].fillna(df["MonthlyIncome"].median())
df["JobSatisfaction"] = df["JobSatisfaction"].fillna(df["JobSatisfaction"].median())
df["YearsAtCompany"] = df["YearsAtCompany"].fillna(df["YearsAtCompany"].median())
df["MonthlyIncome"] = df["MonthlyIncome"].round(0).astype(int)
df["JobSatisfaction"] = df["JobSatisfaction"].round(0).astype(int)
df["YearsAtCompany"] = df["YearsAtCompany"].round(0).astype(int)


def salary_band(s):
    if s < 30000:
        return "Low (<30K)"
    if s < 60000:
        return "Mid (30-60K)"
    if s < 100000:
        return "High (60-100K)"
    return "Very High (100K+)"


def age_group(a):
    if a < 25:
        return "18-24"
    if a < 35:
        return "25-34"
    if a < 45:
        return "35-44"
    if a < 55:
        return "45-54"
    return "55+"


def tenure_group(t):
    if t < 2:
        return "0-1 yrs"
    if t < 5:
        return "2-4 yrs"
    if t < 10:
        return "5-9 yrs"
    if t < 20:
        return "10-19 yrs"
    return "20+ yrs"


df["SalaryBand"] = df["MonthlyIncome"].apply(salary_band)
df["AgeGroup"] = df["Age"].apply(age_group)
df["TenureGroup"] = df["YearsAtCompany"].apply(tenure_group)
df["AttritionNum"] = (df["Attrition"] == "Yes").astype(int)

total = len(df)
attr_count = df["AttritionNum"].sum()
attr_rate = attr_count / total * 100
print(f"   Clean: {total} rows | Attrition: {attr_count} ({attr_rate:.1f}%)")

# ══════════════════════════════════════════════
# 3. CHARTS
# ══════════════════════════════════════════════
print("▶ Step 3/5 — Generating 4 charts...")

PALETTE = ["#d4380d", "#0958d9", "#389e0d", "#d46b08", "#531dab", "#aaa"]
BG, INK = "#FAFAF8", "#1a1410"
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.facecolor": BG,
    "figure.facecolor": BG,
    "axes.grid": True,
    "grid.color": "#e8e4de",
    "grid.linewidth": 0.6,
})

# Chart 1: Department
fig1, ax = plt.subplots(figsize=(9, 4))
dept_attr = df.groupby("Department")["AttritionNum"].agg(["sum", "count"])
dept_attr["rate"] = dept_attr["sum"] / dept_attr["count"] * 100
dept_attr = dept_attr.sort_values("rate", ascending=True)
bars = ax.barh(dept_attr.index, dept_attr["rate"],
               color=[PALETTE[0] if r == dept_attr["rate"].max() else PALETTE[1]
                      for r in dept_attr["rate"]], height=0.5)
ax.set_xlabel("Attrition Rate (%)", fontsize=9)
ax.set_title("Attrition Rate by Department", fontsize=13, fontweight="bold", color=INK, pad=12)
for bar, val in zip(bars, dept_attr["rate"]):
    ax.text(val + 0.3, bar.get_y() + bar.get_height() / 2,
            f"{val:.1f}%", va="center", fontsize=9, fontweight="bold")
ax.axvline(x=attr_rate, color="gray", linestyle="--", linewidth=1, alpha=0.6)
plt.tight_layout()
c1 = OUT / "chart1_dept_attrition.png"
fig1.savefig(c1, dpi=150, bbox_inches="tight")
plt.close()

# Chart 2: Age
fig2, ax = plt.subplots(figsize=(9, 4))
age_order = ["18-24", "25-34", "35-44", "45-54", "55+"]
age_attr = df.groupby("AgeGroup")["AttritionNum"].agg(["sum", "count"])
age_attr["rate"] = age_attr["sum"] / age_attr["count"] * 100
age_attr = age_attr.reindex([a for a in age_order if a in age_attr.index])
colors_age = [PALETTE[0] if v == age_attr["rate"].max() else "#aac4e8" for v in age_attr["rate"]]
bars2 = ax.bar(age_attr.index, age_attr["rate"], color=colors_age, width=0.5)
ax.set_ylabel("Attrition Rate (%)", fontsize=9)
ax.set_title("Attrition Rate by Age Group", fontsize=13, fontweight="bold", color=INK, pad=12)
for bar, val in zip(bars2, age_attr["rate"]):
    ax.text(bar.get_x() + bar.get_width() / 2, val + 0.4,
            f"{val:.1f}%", ha="center", fontsize=9, fontweight="bold")
ax.axhline(y=attr_rate, color="gray", linestyle="--", linewidth=1, alpha=0.6)
plt.tight_layout()
c2 = OUT / "chart2_age_attrition.png"
fig2.savefig(c2, dpi=150, bbox_inches="tight")
plt.close()

# Chart 3: Salary
fig3, ax = plt.subplots(figsize=(9, 4))
sal_order = ["Low (<30K)", "Mid (30-60K)", "High (60-100K)", "Very High (100K+)"]
sal_attr = df.groupby("SalaryBand")["AttritionNum"].agg(["sum", "count"])
sal_attr["rate"] = sal_attr["sum"] / sal_attr["count"] * 100
sal_attr = sal_attr.reindex([s for s in sal_order if s in sal_attr.index])
colors_sal = [PALETTE[0] if v == sal_attr["rate"].max() else PALETTE[1] for v in sal_attr["rate"]]
bars3 = ax.bar(sal_attr.index, sal_attr["rate"], color=colors_sal, width=0.5)
ax.set_ylabel("Attrition Rate (%)", fontsize=9)
ax.set_title("Attrition Rate by Salary Band", fontsize=13, fontweight="bold", color=INK, pad=12)
ax.tick_params(axis="x", rotation=15)
for bar, val in zip(bars3, sal_attr["rate"]):
    ax.text(bar.get_x() + bar.get_width() / 2, val + 0.4,
            f"{val:.1f}%", ha="center", fontsize=9, fontweight="bold")
plt.tight_layout()
c3 = OUT / "chart3_salary_attrition.png"
fig3.savefig(c3, dpi=150, bbox_inches="tight")
plt.close()

# Chart 4: Tenure
fig4, ax = plt.subplots(figsize=(9, 4))
ten_order = ["0-1 yrs", "2-4 yrs", "5-9 yrs", "10-19 yrs", "20+ yrs"]
ten_attr = df.groupby("TenureGroup")["AttritionNum"].agg(["sum", "count"])
ten_attr["rate"] = ten_attr["sum"] / ten_attr["count"] * 100
ten_attr = ten_attr.reindex([t for t in ten_order if t in ten_attr.index])
ax.plot(ten_attr.index, ten_attr["rate"], color=PALETTE[0],
        linewidth=2.5, marker="o", markersize=8, zorder=3)
ax.fill_between(range(len(ten_attr)), ten_attr["rate"], alpha=0.10, color=PALETTE[0])
for i, val in enumerate(ten_attr["rate"]):
    ax.text(i, val + 0.5, f"{val:.1f}%", ha="center", fontsize=9, fontweight="bold")
ax.set_xticks(range(len(ten_attr)))
ax.set_xticklabels(ten_attr.index)
ax.set_ylabel("Attrition Rate (%)", fontsize=9)
ax.set_title("Attrition Rate by Tenure", fontsize=13, fontweight="bold", color=INK, pad=12)
plt.tight_layout()
c4 = OUT / "chart4_tenure_attrition.png"
fig4.savefig(c4, dpi=150, bbox_inches="tight")
plt.close()
print("   4 charts saved")

# ══════════════════════════════════════════════
# 4. SUMMARY
# ══════════════════════════════════════════════
print("▶ Step 4/5 — Generating executive summary...")
high_dept = dept_attr["rate"].idxmax()
high_age = age_attr["rate"].idxmax()
low_sal_rate = sal_attr["rate"].iloc[0]
high_sal_rate = sal_attr["rate"].iloc[-1]
high_ten = ten_attr["rate"].idxmax()
ot_attr = df[df["OverTime"] == "Yes"]["AttritionNum"].mean() * 100
no_ot_attr = df[df["OverTime"] == "No"]["AttritionNum"].mean() * 100
sal_multiple = low_sal_rate / high_sal_rate if high_sal_rate else 0

insights = (
    f"The organization is experiencing an overall attrition rate of {attr_rate:.1f}%, "
    f"with {attr_count} employees leaving out of {total}. The {high_dept} department shows "
    f"the highest attrition at {dept_attr['rate'].max():.1f}%, significantly above the company "
    "average. Immediate intervention through targeted retention programs, manager training, and "
    "role satisfaction surveys is recommended for this department.\n\n"
    f"Compensation remains a critical driver of attrition. Employees in the lowest salary band "
    f"(<30K) leave at a rate of {low_sal_rate:.1f}% — nearly {sal_multiple:.1f}x higher than the "
    f"highest earners. Additionally, employees working overtime leave at {ot_attr:.1f}% versus "
    f"{no_ot_attr:.1f}% for those who don't, indicating workload imbalance. A salary review for "
    "low-band employees and overtime policy reform should be prioritized in the next budget cycle.\n\n"
    f"Tenure analysis reveals that employees in the {high_ten} bracket have the highest attrition "
    f"at {ten_attr['rate'].max():.1f}%, suggesting a critical retention window. The {high_age} age "
    "group also shows elevated attrition, pointing to career growth concerns among early-career "
    "employees. Implementing structured career development plans, mentorship programs, and "
    "promotion timelines for employees in their first two years could significantly reduce "
    "overall attrition."
)

# ══════════════════════════════════════════════
# 5. EXCEL EXPORT
# ══════════════════════════════════════════════
print("▶ Step 5/5 — Building polished Excel...")

excel_path = OUT / "HR_Attrition_Analysis.xlsx"

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Raw Data", index=False)

    summary_df = pd.DataFrame([
        ["Total Employees", total],
        ["Employees Left", int(attr_count)],
        ["Attrition Rate", f"{attr_rate:.1f}%"],
        ["Active Employees", total - int(attr_count)],
        ["Avg Monthly Income", f"{df['MonthlyIncome'].mean():,.0f}"],
        ["Avg Age", f"{df['Age'].mean():.1f} yrs"],
        ["Avg Tenure", f"{df['YearsAtCompany'].mean():.1f} yrs"],
        ["OT Attrition Rate", f"{ot_attr:.1f}%"],
        ["Non-OT Attrition Rate", f"{no_ot_attr:.1f}%"],
        ["Highest Risk Dept", high_dept],
        ["Highest Risk Age", high_age],
        ["Highest Risk Tenure", high_ten],
    ], columns=["Metric", "Value"])
    summary_df.to_excel(writer, sheet_name="KPI Summary", index=False)

    dept_full = df.groupby("Department").agg(
        Total=("AttritionNum", "count"),
        Left=("AttritionNum", "sum"),
        Avg_Salary=("MonthlyIncome", "mean"),
        Avg_Tenure=("YearsAtCompany", "mean"),
    ).round(1).reset_index()
    dept_full["Attrition_Rate"] = (dept_full["Left"] / dept_full["Total"] * 100).round(1)
    dept_full.to_excel(writer, sheet_name="Dept Analysis", index=False)

    pd.DataFrame({"Executive Summary": [insights]}).to_excel(
        writer, sheet_name="Executive Summary", index=False
    )

wb = load_workbook(excel_path)
ws = wb.create_sheet("Dashboard", 0)

DARK_BLUE = PatternFill("solid", fgColor="1B4F8A")
ORANGE = PatternFill("solid", fgColor="D4380D")
title_font = Font(color="FFFFFF", bold=True, name="Arial", size=16)

for col, w in [(1, 3), (2, 22), (3, 18), (4, 18), (5, 18), (6, 18), (7, 3)]:
    ws.column_dimensions[get_column_letter(col)].width = w

ws.merge_cells("B1:F2")
ws["B1"] = "  HR ATTRITION ANALYSIS DASHBOARD"
ws["B1"].fill = DARK_BLUE
ws["B1"].font = title_font
ws["B1"].alignment = Alignment(vertical="center", horizontal="left")
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22

kpis = [
    ("TOTAL EMPLOYEES", total, DARK_BLUE),
    ("ATTRITION RATE", f"{attr_rate:.1f}%", ORANGE),
    ("EMPLOYEES LEFT", int(attr_count), PatternFill("solid", fgColor="531DAB")),
    ("AVG TENURE", f"{df['YearsAtCompany'].mean():.1f} yrs",
     PatternFill("solid", fgColor="389E0D")),
]
for col, (label, value, fill) in zip([2, 3, 4, 5], kpis):
    cl = get_column_letter(col)
    ws[f"{cl}4"] = label
    ws[f"{cl}4"].fill = fill
    ws[f"{cl}4"].font = Font(color="FFFFFF", bold=True, name="Arial", size=8)
    ws[f"{cl}4"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{cl}5"] = value
    ws[f"{cl}5"].fill = fill
    ws[f"{cl}5"].font = Font(color="FFFFFF", bold=True, name="Arial", size=14)
    ws[f"{cl}5"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{cl}6"].fill = fill

ws.row_dimensions[4].height = 14
ws.row_dimensions[5].height = 26
ws.row_dimensions[6].height = 14

ws.merge_cells("B8:F8")
ws["B8"] = "  ATTRITION ANALYSIS — 4 DIMENSIONS"
ws["B8"].fill = PatternFill("solid", fgColor="2E4057")
ws["B8"].font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
ws["B8"].alignment = Alignment(vertical="center", horizontal="left")
ws.row_dimensions[8].height = 18

for path, anchor in [(c1, "B9"), (c2, "B29"), (c3, "B49"), (c4, "B69")]:
    img = XLImage(str(path))
    img.width, img.height = 430, 190
    ws.add_image(img, anchor)

ins_row = 90
ws.merge_cells(f"B{ins_row}:F{ins_row}")
ws[f"B{ins_row}"] = "  EXECUTIVE SUMMARY"
ws[f"B{ins_row}"].fill = DARK_BLUE
ws[f"B{ins_row}"].font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
ws[f"B{ins_row}"].alignment = Alignment(vertical="center")
ws.row_dimensions[ins_row].height = 18

for j, para in enumerate(insights.split("\n\n")[:3], 1):
    r = ins_row + j
    ws.merge_cells(f"B{r}:F{r}")
    ws[f"B{r}"] = para.strip()
    ws[f"B{r}"].font = Font(name="Arial", size=9, color="333333")
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 60

wb.save(excel_path)
print(f"   Excel Dashboard saved → {excel_path}")

print(f"""
╔══════════════════════════════════════════════╗
║      HR ATTRITION PIPELINE — COMPLETE        ║
╠══════════════════════════════════════════════╣
║  Output folder : {str(OUT.relative_to(PROJECT_ROOT)):<28} ║
║  Excel report  : HR_Attrition_Analysis.xlsx  ║
║  Charts        : chart1 / chart2 / chart3 / chart4 ║
╚══════════════════════════════════════════════╝
""")
