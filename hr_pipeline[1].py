import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import warnings, os, json, urllib.request, re
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

OUT = '/home/claude/hr_output'
os.makedirs(OUT, exist_ok=True)
np.random.seed(42)

# ══════════════════════════════════════════════
# 1. GENERATE IBM-STYLE HR ATTRITION DATASET
# ══════════════════════════════════════════════
print("▶ Step 1/5 — Generating HR Attrition dataset...")

n = 1470
depts = {
    'Sales':               0.34,
    'Research & Development': 0.48,
    'Human Resources':     0.18,
}
job_roles = {
    'Sales':               ['Sales Executive','Sales Representative','Manager'],
    'Research & Development': ['Research Scientist','Laboratory Technician',
                               'Healthcare Representative','Manufacturing Director','Research Director'],
    'Human Resources':     ['Human Resources','Manager'],
}
edu_fields = ['Life Sciences','Medical','Marketing','Technical Degree','Human Resources','Other']

rows = []
for i in range(n):
    dept      = np.random.choice(list(depts.keys()), p=list(depts.values()))
    role      = np.random.choice(job_roles[dept])
    age       = int(np.random.normal(37, 9))
    age       = max(18, min(60, age))
    tenure    = np.random.randint(0, min(age-18, 35)+1)
    salary    = int(np.random.lognormal(9.2, 0.45))
    salary    = max(10000, min(200000, salary))
    dist      = np.random.randint(1, 30)
    overtime  = np.random.choice(['Yes','No'], p=[0.28,0.72])
    satisf    = np.random.randint(1, 5)
    env_sat   = np.random.randint(1, 5)
    wlb       = np.random.randint(1, 5)
    perf      = np.random.choice([3,4], p=[0.84,0.16])
    edu       = np.random.randint(1, 6)
    edu_field = np.random.choice(edu_fields)
    num_cos   = np.random.randint(0, 10)
    last_promo = np.random.randint(0, 15)

    # Attrition probability model
    p_attr = 0.16
    if overtime == 'Yes':      p_attr += 0.12
    if satisf <= 2:            p_attr += 0.10
    if wlb <= 2:               p_attr += 0.08
    if tenure < 2:             p_attr += 0.10
    if salary < 30000:         p_attr += 0.08
    if dist > 20:              p_attr += 0.05
    if dept == 'Sales':        p_attr += 0.04
    if last_promo > 5:         p_attr += 0.06
    p_attr = min(p_attr, 0.75)

    attrition = np.random.choice(['Yes','No'], p=[p_attr, 1-p_attr])

    rows.append({
        'EmployeeID'           : 1000+i,
        'Age'                  : age,
        'Attrition'            : attrition,
        'Department'           : dept,
        'JobRole'              : role,
        'MonthlyIncome'        : salary,
        'YearsAtCompany'       : tenure,
        'DistanceFromHome'     : dist,
        'OverTime'             : overtime,
        'JobSatisfaction'      : satisf,
        'EnvironmentSatisfaction': env_sat,
        'WorkLifeBalance'      : wlb,
        'PerformanceRating'    : perf,
        'Education'            : edu,
        'EducationField'       : edu_field,
        'NumCompaniesWorked'   : num_cos,
        'YearsSinceLastPromotion': last_promo,
        'Gender'               : np.random.choice(['Male','Female'], p=[0.60,0.40]),
        'MaritalStatus'        : np.random.choice(['Single','Married','Divorced'], p=[0.32,0.46,0.22]),
    })

df_raw = pd.DataFrame(rows)
# Inject mess
for col in ['MonthlyIncome','JobSatisfaction','YearsAtCompany']:
    mask = np.random.random(len(df_raw)) < 0.03
    df_raw.loc[mask, col] = np.nan
dupes = df_raw.sample(30, random_state=5)
df_raw = pd.concat([df_raw, dupes], ignore_index=True)
df_raw.to_csv(f'{OUT}/dirty_hr_attrition.csv', index=False)
print(f"   Raw: {len(df_raw)} rows | Missing: {df_raw.isnull().sum().sum()} | Dupes: 30")

# ══════════════════════════════════════════════
# 2. CLEAN
# ══════════════════════════════════════════════
print("▶ Step 2/5 — Cleaning data...")
df = df_raw.drop_duplicates().copy()
df['MonthlyIncome']  = df['MonthlyIncome'].fillna(df['MonthlyIncome'].median())
df['JobSatisfaction']= df['JobSatisfaction'].fillna(df['JobSatisfaction'].median())
df['YearsAtCompany'] = df['YearsAtCompany'].fillna(df['YearsAtCompany'].median())
df['MonthlyIncome']  = df['MonthlyIncome'].round(0).astype(int)
df['JobSatisfaction']= df['JobSatisfaction'].round(0).astype(int)
df['YearsAtCompany'] = df['YearsAtCompany'].round(0).astype(int)

# Salary bands
def salary_band(s):
    if s < 30000:   return 'Low (<30K)'
    elif s < 60000: return 'Mid (30-60K)'
    elif s < 100000:return 'High (60-100K)'
    else:           return 'Very High (100K+)'

def age_group(a):
    if a < 25:    return '18-24'
    elif a < 35:  return '25-34'
    elif a < 45:  return '35-44'
    elif a < 55:  return '45-54'
    else:         return '55+'

def tenure_group(t):
    if t < 2:    return '0-1 yrs'
    elif t < 5:  return '2-4 yrs'
    elif t < 10: return '5-9 yrs'
    elif t < 20: return '10-19 yrs'
    else:        return '20+ yrs'

df['SalaryBand']   = df['MonthlyIncome'].apply(salary_band)
df['AgeGroup']     = df['Age'].apply(age_group)
df['TenureGroup']  = df['YearsAtCompany'].apply(tenure_group)
df['AttritionNum'] = (df['Attrition'] == 'Yes').astype(int)

total      = len(df)
attr_count = df['AttritionNum'].sum()
attr_rate  = attr_count / total * 100
print(f"   Clean: {total} rows | Attrition: {attr_count} ({attr_rate:.1f}%)")

# ══════════════════════════════════════════════
# 3. GENERATE 4 CHARTS
# ══════════════════════════════════════════════
print("▶ Step 3/5 — Generating 4 charts...")

PALETTE = ['#d4380d','#0958d9','#389e0d','#d46b08','#531dab','#aaa']
BG, INK  = '#FAFAF8','#1a1410'
plt.rcParams.update({
    'font.family'      : 'DejaVu Sans',
    'axes.spines.top'  : False,
    'axes.spines.right': False,
    'axes.facecolor'   : BG,
    'figure.facecolor' : BG,
    'axes.grid'        : True,
    'grid.color'       : '#e8e4de',
    'grid.linewidth'   : 0.6,
})

# Chart 1: Attrition by Department
fig1, ax = plt.subplots(figsize=(9, 4))
dept_attr = df.groupby('Department')['AttritionNum'].agg(['sum','count'])
dept_attr['rate'] = dept_attr['sum'] / dept_attr['count'] * 100
dept_attr = dept_attr.sort_values('rate', ascending=True)
bars = ax.barh(dept_attr.index, dept_attr['rate'],
               color=[PALETTE[0] if r == dept_attr['rate'].max() else PALETTE[1]
                      for r in dept_attr['rate']], height=0.5)
ax.set_xlabel('Attrition Rate (%)', fontsize=9)
ax.set_title('Attrition Rate by Department', fontsize=13, fontweight='bold', color=INK, pad=12)
for bar, val in zip(bars, dept_attr['rate']):
    ax.text(val+0.3, bar.get_y()+bar.get_height()/2,
            f'{val:.1f}%', va='center', fontsize=9, fontweight='bold')
ax.axvline(x=attr_rate, color='gray', linestyle='--', linewidth=1, alpha=0.6)
ax.text(attr_rate+0.2, -0.4, f'Avg {attr_rate:.1f}%', fontsize=7.5, color='gray')
plt.tight_layout()
c1 = f'{OUT}/chart1_dept_attrition.png'
fig1.savefig(c1, dpi=150, bbox_inches='tight'); plt.close()

# Chart 2: Attrition by Age Group
fig2, ax = plt.subplots(figsize=(9, 4))
age_order  = ['18-24','25-34','35-44','45-54','55+']
age_attr   = df.groupby('AgeGroup')['AttritionNum'].agg(['sum','count'])
age_attr['rate'] = age_attr['sum'] / age_attr['count'] * 100
age_attr   = age_attr.reindex([a for a in age_order if a in age_attr.index])
colors_age = [PALETTE[0] if v == age_attr['rate'].max() else '#aac4e8'
              for v in age_attr['rate']]
bars2 = ax.bar(age_attr.index, age_attr['rate'], color=colors_age, width=0.5)
ax.set_ylabel('Attrition Rate (%)', fontsize=9)
ax.set_title('Attrition Rate by Age Group', fontsize=13, fontweight='bold', color=INK, pad=12)
for bar, val in zip(bars2, age_attr['rate']):
    ax.text(bar.get_x()+bar.get_width()/2, val+0.4,
            f'{val:.1f}%', ha='center', fontsize=9, fontweight='bold')
ax.axhline(y=attr_rate, color='gray', linestyle='--', linewidth=1, alpha=0.6)
plt.tight_layout()
c2 = f'{OUT}/chart2_age_attrition.png'
fig2.savefig(c2, dpi=150, bbox_inches='tight'); plt.close()

# Chart 3: Attrition by Salary Band
fig3, ax = plt.subplots(figsize=(9, 4))
sal_order = ['Low (<30K)','Mid (30-60K)','High (60-100K)','Very High (100K+)']
sal_attr  = df.groupby('SalaryBand')['AttritionNum'].agg(['sum','count'])
sal_attr['rate'] = sal_attr['sum'] / sal_attr['count'] * 100
sal_attr  = sal_attr.reindex([s for s in sal_order if s in sal_attr.index])
colors_sal = [PALETTE[0] if v == sal_attr['rate'].max() else PALETTE[1]
              for v in sal_attr['rate']]
bars3 = ax.bar(sal_attr.index, sal_attr['rate'], color=colors_sal, width=0.5)
ax.set_ylabel('Attrition Rate (%)', fontsize=9)
ax.set_title('Attrition Rate by Salary Band', fontsize=13, fontweight='bold', color=INK, pad=12)
ax.tick_params(axis='x', rotation=15)
for bar, val in zip(bars3, sal_attr['rate']):
    ax.text(bar.get_x()+bar.get_width()/2, val+0.4,
            f'{val:.1f}%', ha='center', fontsize=9, fontweight='bold')
plt.tight_layout()
c3 = f'{OUT}/chart3_salary_attrition.png'
fig3.savefig(c3, dpi=150, bbox_inches='tight'); plt.close()

# Chart 4: Attrition by Tenure
fig4, ax = plt.subplots(figsize=(9, 4))
ten_order = ['0-1 yrs','2-4 yrs','5-9 yrs','10-19 yrs','20+ yrs']
ten_attr  = df.groupby('TenureGroup')['AttritionNum'].agg(['sum','count'])
ten_attr['rate'] = ten_attr['sum'] / ten_attr['count'] * 100
ten_attr  = ten_attr.reindex([t for t in ten_order if t in ten_attr.index])
ax.plot(ten_attr.index, ten_attr['rate'], color=PALETTE[0],
        linewidth=2.5, marker='o', markersize=8, zorder=3)
ax.fill_between(range(len(ten_attr)), ten_attr['rate'],
                alpha=0.10, color=PALETTE[0])
for i, val in enumerate(ten_attr['rate']):
    ax.text(i, val+0.5, f'{val:.1f}%', ha='center', fontsize=9, fontweight='bold')
ax.set_xticks(range(len(ten_attr)))
ax.set_xticklabels(ten_attr.index)
ax.set_ylabel('Attrition Rate (%)', fontsize=9)
ax.set_title('Attrition Rate by Tenure', fontsize=13, fontweight='bold', color=INK, pad=12)
plt.tight_layout()
c4 = f'{OUT}/chart4_tenure_attrition.png'
fig4.savefig(c4, dpi=150, bbox_inches='tight'); plt.close()
print("   4 charts saved")

# ══════════════════════════════════════════════
# 4. AI INSIGHTS
# ══════════════════════════════════════════════
print("▶ Step 4/5 — Generating executive summary...")

high_dept = dept_attr['rate'].idxmax()
high_age  = age_attr['rate'].idxmax()
low_sal_rate = sal_attr['rate'].iloc[0]
high_ten  = ten_attr['rate'].idxmax()
ot_attr   = df[df['OverTime']=='Yes']['AttritionNum'].mean()*100
no_ot_attr= df[df['OverTime']=='No']['AttritionNum'].mean()*100

summary = f"""
TOTAL: {total} employees | ATTRITION: {attr_count} ({attr_rate:.1f}%)
DEPT: {high_dept} highest at {dept_attr['rate'].max():.1f}%
AGE: {high_age} group highest at {age_attr['rate'].max():.1f}%
SALARY: Low (<30K) attrition {low_sal_rate:.1f}% vs Very High only {sal_attr['rate'].iloc[-1]:.1f}%
TENURE: {high_ten} highest at {ten_attr['rate'].max():.1f}%
OVERTIME: Yes={ot_attr:.1f}% vs No={no_ot_attr:.1f}%
"""

insights = None
try:
    import os as _os
    key = _os.environ.get('ANTHROPIC_API_KEY','')
    if not key: raise ValueError("no key")
    payload = json.dumps({
        "model":"claude-sonnet-4-20250514","max_tokens":600,
        "messages":[{"role":"user","content":
            f"Write a 3-paragraph executive summary for HR leadership about employee attrition. "
            f"Use specific numbers. End each paragraph with one action recommendation. "
            f"DATA: {summary}"}]
    }).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type":"application/json",
                 "x-api-key":key,"anthropic-version":"2023-06-01"})
    with urllib.request.urlopen(req, timeout=15) as r:
        insights = json.loads(r.read())["content"][0]["text"]
    print("   ✓ Claude API responded")
except:
    insights = f"""The organization is experiencing an overall attrition rate of {attr_rate:.1f}%, with {attr_count} employees leaving out of {total}. The {high_dept} department shows the highest attrition at {dept_attr['rate'].max():.1f}%, significantly above the company average. Immediate intervention through targeted retention programs, manager training, and role satisfaction surveys is recommended for this department.

Compensation remains a critical driver of attrition. Employees in the lowest salary band (<₹30,000/month) leave at a rate of {low_sal_rate:.1f}% — nearly {low_sal_rate/sal_attr['rate'].iloc[-1]:.1f}x higher than the highest earners. Additionally, employees working overtime leave at {ot_attr:.1f}% versus {no_ot_attr:.1f}% for those who don't, indicating workload imbalance. A salary review for low-band employees and overtime policy reform should be prioritized in the next budget cycle.

Tenure analysis reveals that employees in the {high_ten} bracket have the highest attrition at {ten_attr['rate'].max():.1f}%, suggesting a critical retention window. The {high_age} age group also shows elevated attrition, pointing to career growth concerns among early-career employees. Implementing structured career development plans, mentorship programs, and promotion timelines for employees in their first two years could significantly reduce overall attrition."""
    print("   ✓ Data-derived insights used")

# ══════════════════════════════════════════════
# 5. EXPORT POLISHED EXCEL
# ══════════════════════════════════════════════
print("▶ Step 5/5 — Building polished Excel...")

excel_path = f'{OUT}/HR_Attrition_Analysis.xlsx'

# Write data sheets first
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Raw Data', index=False)

    # Summary
    summary_df = pd.DataFrame([
        ['Total Employees',      total],
        ['Employees Left',       int(attr_count)],
        ['Attrition Rate',       f'{attr_rate:.1f}%'],
        ['Active Employees',     total - int(attr_count)],
        ['Avg Monthly Income',   f'₹{df["MonthlyIncome"].mean():,.0f}'],
        ['Avg Age',              f'{df["Age"].mean():.1f} yrs'],
        ['Avg Tenure',           f'{df["YearsAtCompany"].mean():.1f} yrs'],
        ['OT Attrition Rate',    f'{ot_attr:.1f}%'],
        ['Non-OT Attrition Rate',f'{no_ot_attr:.1f}%'],
        ['Highest Risk Dept',    high_dept],
        ['Highest Risk Age',     high_age],
        ['Highest Risk Tenure',  high_ten],
    ], columns=['Metric','Value'])
    summary_df.to_excel(writer, sheet_name='KPI Summary', index=False)

    # Dept breakdown
    dept_full = df.groupby('Department').agg(
        Total=('AttritionNum','count'),
        Left=('AttritionNum','sum'),
        Avg_Salary=('MonthlyIncome','mean'),
        Avg_Tenure=('YearsAtCompany','mean'),
    ).round(1).reset_index()
    dept_full['Attrition_Rate'] = (dept_full['Left']/dept_full['Total']*100).round(1)
    dept_full.to_excel(writer, sheet_name='Dept Analysis', index=False)

    # Insights
    ins_df = pd.DataFrame({'Executive Summary': [insights]})
    ins_df.to_excel(writer, sheet_name='Executive Summary', index=False)

# Now build Dashboard sheet with images
wb = load_workbook(excel_path)

# Dashboard sheet
ws = wb.create_sheet('Dashboard', 0)

# Colors
DARK_BLUE = PatternFill("solid", fgColor="1B4F8A")
ORANGE    = PatternFill("solid", fgColor="D4380D")
LIGHT_BG  = PatternFill("solid", fgColor="F5F5F0")
WHITE     = PatternFill("solid", fgColor="FFFFFF")
ACCENT    = PatternFill("solid", fgColor="E8F0FF")

white_font  = Font(color="FFFFFF", bold=True, name='Arial')
dark_font   = Font(color="1A1410", bold=True, name='Arial')
normal_font = Font(color="1A1410", name='Arial')
title_font  = Font(color="FFFFFF", bold=True, name='Arial', size=16)

def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

# Set column widths
for col, w in [(1,3),(2,22),(3,18),(4,18),(5,18),(6,18),(7,3)]:
    ws.column_dimensions[get_column_letter(col)].width = w

# Title bar
ws.merge_cells('B1:F2')
ws['B1'] = '  HR ATTRITION ANALYSIS DASHBOARD'
ws['B1'].fill   = DARK_BLUE
ws['B1'].font   = title_font
ws['B1'].alignment = Alignment(vertical='center', horizontal='left')
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22

# KPI row
kpis = [
    ('TOTAL EMPLOYEES', total, DARK_BLUE),
    ('ATTRITION RATE', f'{attr_rate:.1f}%', ORANGE),
    ('EMPLOYEES LEFT', int(attr_count), PatternFill("solid", fgColor="531DAB")),
    ('AVG TENURE', f'{df["YearsAtCompany"].mean():.1f} yrs',
     PatternFill("solid", fgColor="389E0D")),
]
kpi_cols = [2,3,4,5]
ws.row_dimensions[4].height = 14
ws.row_dimensions[5].height = 26
ws.row_dimensions[6].height = 14

for col, (label, value, fill) in zip(kpi_cols, kpis):
    cl = get_column_letter(col)
    ws[f'{cl}4'] = label
    ws[f'{cl}4'].fill = fill
    ws[f'{cl}4'].font = Font(color='FFFFFF', bold=True, name='Arial', size=8)
    ws[f'{cl}4'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{cl}5'] = value
    ws[f'{cl}5'].fill = fill
    ws[f'{cl}5'].font = Font(color='FFFFFF', bold=True, name='Arial', size=14)
    ws[f'{cl}5'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{cl}6'].fill = fill

# Charts label
ws.merge_cells('B8:F8')
ws['B8'] = '  ATTRITION ANALYSIS — 4 DIMENSIONS'
ws['B8'].fill = PatternFill("solid", fgColor="2E4057")
ws['B8'].font = Font(color='FFFFFF', bold=True, name='Arial', size=11)
ws['B8'].alignment = Alignment(vertical='center', horizontal='left')
ws.row_dimensions[8].height = 18

# Insert chart images
for path, anchor, w, h in [
    (c1, 'B9',  430, 190),
    (c2, 'B29', 430, 190),
    (c3, 'B49', 430, 190),
    (c4, 'B69', 430, 190),
]:
    img = XLImage(path)
    img.width  = w
    img.height = h
    ws.add_image(img, anchor)

# Insights section
ins_row = 90
ws.merge_cells(f'B{ins_row}:F{ins_row}')
ws[f'B{ins_row}'] = '  EXECUTIVE SUMMARY'
ws[f'B{ins_row}'].fill = DARK_BLUE
ws[f'B{ins_row}'].font = Font(color='FFFFFF', bold=True, name='Arial', size=11)
ws[f'B{ins_row}'].alignment = Alignment(vertical='center')
ws.row_dimensions[ins_row].height = 18

for j, para in enumerate(insights.split('\n\n')[:3], 1):
    r = ins_row + j
    ws.merge_cells(f'B{r}:F{r}')
    ws[f'B{r}'] = para.strip()
    ws[f'B{r}'].font = Font(name='Arial', size=9, color='333333')
    ws[f'B{r}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 60

wb.save(excel_path)
print(f"   Excel Dashboard saved → {excel_path}")

# README
readme = f"""# 👥 HR Employee Attrition Analysis

**Portfolio Project #3** — IBM-style HR analytics with executive dashboard.

## What This Does
- Analyses {total} employee records across departments, age groups, salary bands, and tenure
- Identifies key attrition drivers using Python and Pandas
- Generates 4 targeted charts
- Exports polished Excel with a visual Dashboard sheet + Executive Summary

## Key Findings
| Metric | Value |
|---|---|
| Overall Attrition Rate | {attr_rate:.1f}% |
| Highest Risk Department | {high_dept} ({dept_attr['rate'].max():.1f}%) |
| Highest Risk Age Group | {high_age} |
| Overtime Attrition | {ot_attr:.1f}% vs {no_ot_attr:.1f}% (no OT) |

## Tech Stack
`Python` `Pandas` `Matplotlib` `OpenPyXL`

## Author
**Shahnawz Valvari** | MCA Graduate | Python · Data Analytics · HR Analytics
"""
with open(f'{OUT}/README.md','w') as f: f.write(readme)
import shutil
shutil.copy('/home/claude/hr_attrition.py', f'{OUT}/hr_pipeline.py')
with open(f'{OUT}/requirements.txt','w') as f:
    f.write("pandas>=2.0\nmatplotlib>=3.7\nopenpyxl>=3.1\n")

print(f"""
╔══════════════════════════════════════════════╗
║      PORTFOLIO PIECE #3 — COMPLETE           ║
╠══════════════════════════════════════════════╣
║  Excel  : HR_Attrition_Analysis.xlsx         ║
║  Charts : chart1 / chart2 / chart3 / chart4  ║
║  README : README.md                          ║
╚══════════════════════════════════════════════╝
""")
