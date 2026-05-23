"""HR Attrition Analytics — Streamlit dashboard.

Interactive version of the HR attrition pipeline. Loads / generates the dataset,
shows KPIs, four analysis charts, an executive summary, and download buttons
for the Excel report and raw CSV.
"""
from __future__ import annotations

import io

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Page config + theme
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="HR Attrition Analytics",
    page_icon="👥",
    layout="wide",
    initial_sidebar_state="expanded",
)

PALETTE = ["#d4380d", "#0958d9", "#389e0d", "#d46b08", "#531dab", "#aaa"]
BG, INK = "#FAFAF8", "#1a1410"

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.facecolor": BG,
    "figure.facecolor": BG,
    "axes.grid": True,
    "grid.color": "#e8e4de",
    "grid.linewidth": 0.6,
})


# ──────────────────────────────────────────────────────────────────────────────
# Data generation + cleaning (cached)
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def generate_dataset(n: int = 1470, seed: int = 42) -> pd.DataFrame:
    """Generate an IBM-style HR attrition dataset with realistic noise."""
    rng = np.random.default_rng(seed)
    depts = {"Sales": 0.34, "Research & Development": 0.48, "Human Resources": 0.18}
    job_roles = {
        "Sales": ["Sales Executive", "Sales Representative", "Manager"],
        "Research & Development": [
            "Research Scientist", "Laboratory Technician", "Healthcare Representative",
            "Manufacturing Director", "Research Director",
        ],
        "Human Resources": ["Human Resources", "Manager"],
    }
    edu_fields = ["Life Sciences", "Medical", "Marketing", "Technical Degree", "Human Resources", "Other"]

    rows = []
    for i in range(n):
        dept = rng.choice(list(depts.keys()), p=list(depts.values()))
        role = rng.choice(job_roles[dept])
        age = int(np.clip(rng.normal(37, 9), 18, 60))
        tenure = int(rng.integers(0, min(age - 18, 35) + 1))
        salary = int(np.clip(rng.lognormal(9.2, 0.45), 10000, 200000))
        dist = int(rng.integers(1, 30))
        overtime = rng.choice(["Yes", "No"], p=[0.28, 0.72])
        satisf = int(rng.integers(1, 5))
        env_sat = int(rng.integers(1, 5))
        wlb = int(rng.integers(1, 5))
        perf = int(rng.choice([3, 4], p=[0.84, 0.16]))
        edu = int(rng.integers(1, 6))
        edu_field = rng.choice(edu_fields)
        num_cos = int(rng.integers(0, 10))
        last_promo = int(rng.integers(0, 15))

        # Probabilistic attrition based on known risk factors
        p_attr = 0.16
        if overtime == "Yes":
            p_attr += 0.12
        if satisf <= 2:
            p_attr += 0.10
        if wlb <= 2:
            p_attr += 0.08
        if tenure < 2:
            p_attr += 0.10
        if salary < 30000:
            p_attr += 0.08
        if dist > 20:
            p_attr += 0.05
        if dept == "Sales":
            p_attr += 0.04
        if last_promo > 5:
            p_attr += 0.06
        p_attr = min(p_attr, 0.75)

        attrition = rng.choice(["Yes", "No"], p=[p_attr, 1 - p_attr])

        rows.append({
            "EmployeeID": 1000 + i,
            "Age": age,
            "Attrition": attrition,
            "Department": dept,
            "JobRole": role,
            "MonthlyIncome": salary,
            "YearsAtCompany": tenure,
            "DistanceFromHome": dist,
            "OverTime": overtime,
            "JobSatisfaction": satisf,
            "EnvironmentSatisfaction": env_sat,
            "WorkLifeBalance": wlb,
            "PerformanceRating": perf,
            "Education": edu,
            "EducationField": edu_field,
            "NumCompaniesWorked": num_cos,
            "YearsSinceLastPromotion": last_promo,
            "Gender": rng.choice(["Male", "Female"], p=[0.60, 0.40]),
            "MaritalStatus": rng.choice(["Single", "Married", "Divorced"], p=[0.32, 0.46, 0.22]),
        })

    df_raw = pd.DataFrame(rows)

    # Inject realistic data quality issues, then clean them
    for col in ["MonthlyIncome", "JobSatisfaction", "YearsAtCompany"]:
        mask = rng.random(len(df_raw)) < 0.03
        df_raw.loc[mask, col] = np.nan
    dupes = df_raw.sample(30, random_state=5)
    df_raw = pd.concat([df_raw, dupes], ignore_index=True)

    df = df_raw.drop_duplicates().copy()
    df["MonthlyIncome"] = df["MonthlyIncome"].fillna(df["MonthlyIncome"].median())
    df["JobSatisfaction"] = df["JobSatisfaction"].fillna(df["JobSatisfaction"].median())
    df["YearsAtCompany"] = df["YearsAtCompany"].fillna(df["YearsAtCompany"].median())
    df["MonthlyIncome"] = df["MonthlyIncome"].round(0).astype(int)
    df["JobSatisfaction"] = df["JobSatisfaction"].round(0).astype(int)
    df["YearsAtCompany"] = df["YearsAtCompany"].round(0).astype(int)

    df["SalaryBand"] = df["MonthlyIncome"].apply(_salary_band)
    df["AgeGroup"] = df["Age"].apply(_age_group)
    df["TenureGroup"] = df["YearsAtCompany"].apply(_tenure_group)
    df["AttritionNum"] = (df["Attrition"] == "Yes").astype(int)

    return df


def _salary_band(s: float) -> str:
    if s < 30000:
        return "Low (<30K)"
    if s < 60000:
        return "Mid (30-60K)"
    if s < 100000:
        return "High (60-100K)"
    return "Very High (100K+)"


def _age_group(a: int) -> str:
    if a < 25:
        return "18-24"
    if a < 35:
        return "25-34"
    if a < 45:
        return "35-44"
    if a < 55:
        return "45-54"
    return "55+"


def _tenure_group(t: int) -> str:
    if t < 2:
        return "0-1 yrs"
    if t < 5:
        return "2-4 yrs"
    if t < 10:
        return "5-9 yrs"
    if t < 20:
        return "10-19 yrs"
    return "20+ yrs"


# ──────────────────────────────────────────────────────────────────────────────
# Chart builders
# ──────────────────────────────────────────────────────────────────────────────
def chart_dept(df: pd.DataFrame, overall_rate: float):
    fig, ax = plt.subplots(figsize=(9, 4))
    g = df.groupby("Department")["AttritionNum"].agg(["sum", "count"])
    g["rate"] = g["sum"] / g["count"] * 100
    g = g.sort_values("rate")
    colors = [PALETTE[0] if r == g["rate"].max() else PALETTE[1] for r in g["rate"]]
    bars = ax.barh(g.index, g["rate"], color=colors, height=0.5)
    ax.set_xlabel("Attrition Rate (%)", fontsize=9)
    ax.set_title("Attrition Rate by Department", fontsize=13, fontweight="bold", color=INK, pad=12)
    for bar, val in zip(bars, g["rate"]):
        ax.text(val + 0.3, bar.get_y() + bar.get_height() / 2,
                f"{val:.1f}%", va="center", fontsize=9, fontweight="bold")
    ax.axvline(x=overall_rate, color="gray", linestyle="--", linewidth=1, alpha=0.6)
    ax.text(overall_rate + 0.2, -0.4, f"Avg {overall_rate:.1f}%", fontsize=7.5, color="gray")
    plt.tight_layout()
    return fig, g


def chart_age(df: pd.DataFrame, overall_rate: float):
    fig, ax = plt.subplots(figsize=(9, 4))
    order = ["18-24", "25-34", "35-44", "45-54", "55+"]
    g = df.groupby("AgeGroup")["AttritionNum"].agg(["sum", "count"])
    g["rate"] = g["sum"] / g["count"] * 100
    g = g.reindex([a for a in order if a in g.index])
    colors = [PALETTE[0] if v == g["rate"].max() else "#aac4e8" for v in g["rate"]]
    bars = ax.bar(g.index, g["rate"], color=colors, width=0.5)
    ax.set_ylabel("Attrition Rate (%)", fontsize=9)
    ax.set_title("Attrition Rate by Age Group", fontsize=13, fontweight="bold", color=INK, pad=12)
    for bar, val in zip(bars, g["rate"]):
        ax.text(bar.get_x() + bar.get_width() / 2, val + 0.4,
                f"{val:.1f}%", ha="center", fontsize=9, fontweight="bold")
    ax.axhline(y=overall_rate, color="gray", linestyle="--", linewidth=1, alpha=0.6)
    plt.tight_layout()
    return fig, g


def chart_salary(df: pd.DataFrame):
    fig, ax = plt.subplots(figsize=(9, 4))
    order = ["Low (<30K)", "Mid (30-60K)", "High (60-100K)", "Very High (100K+)"]
    g = df.groupby("SalaryBand")["AttritionNum"].agg(["sum", "count"])
    g["rate"] = g["sum"] / g["count"] * 100
    g = g.reindex([s for s in order if s in g.index])
    colors = [PALETTE[0] if v == g["rate"].max() else PALETTE[1] for v in g["rate"]]
    bars = ax.bar(g.index, g["rate"], color=colors, width=0.5)
    ax.set_ylabel("Attrition Rate (%)", fontsize=9)
    ax.set_title("Attrition Rate by Salary Band", fontsize=13, fontweight="bold", color=INK, pad=12)
    ax.tick_params(axis="x", rotation=15)
    for bar, val in zip(bars, g["rate"]):
        ax.text(bar.get_x() + bar.get_width() / 2, val + 0.4,
                f"{val:.1f}%", ha="center", fontsize=9, fontweight="bold")
    plt.tight_layout()
    return fig, g


def chart_tenure(df: pd.DataFrame):
    fig, ax = plt.subplots(figsize=(9, 4))
    order = ["0-1 yrs", "2-4 yrs", "5-9 yrs", "10-19 yrs", "20+ yrs"]
    g = df.groupby("TenureGroup")["AttritionNum"].agg(["sum", "count"])
    g["rate"] = g["sum"] / g["count"] * 100
    g = g.reindex([t for t in order if t in g.index])
    ax.plot(g.index, g["rate"], color=PALETTE[0], linewidth=2.5,
            marker="o", markersize=8, zorder=3)
    ax.fill_between(range(len(g)), g["rate"], alpha=0.10, color=PALETTE[0])
    for i, val in enumerate(g["rate"]):
        ax.text(i, val + 0.5, f"{val:.1f}%", ha="center", fontsize=9, fontweight="bold")
    ax.set_xticks(range(len(g)))
    ax.set_xticklabels(g.index)
    ax.set_ylabel("Attrition Rate (%)", fontsize=9)
    ax.set_title("Attrition Rate by Tenure", fontsize=13, fontweight="bold", color=INK, pad=12)
    plt.tight_layout()
    return fig, g


# ──────────────────────────────────────────────────────────────────────────────
# Insights
# ──────────────────────────────────────────────────────────────────────────────
def build_summary(df: pd.DataFrame, dept_g, age_g, sal_g, ten_g) -> str:
    total = len(df)
    attr_count = int(df["AttritionNum"].sum())
    attr_rate = attr_count / total * 100
    high_dept = dept_g["rate"].idxmax()
    high_age = age_g["rate"].idxmax()
    low_sal_rate = sal_g["rate"].iloc[0]
    high_sal_rate = sal_g["rate"].iloc[-1]
    high_ten = ten_g["rate"].idxmax()
    ot_attr = df[df["OverTime"] == "Yes"]["AttritionNum"].mean() * 100
    no_ot_attr = df[df["OverTime"] == "No"]["AttritionNum"].mean() * 100
    sal_multiple = low_sal_rate / high_sal_rate if high_sal_rate else float("nan")

    return (
        f"The organization is experiencing an overall attrition rate of **{attr_rate:.1f}%**, "
        f"with **{attr_count}** employees leaving out of **{total}**. The **{high_dept}** "
        f"department shows the highest attrition at **{dept_g['rate'].max():.1f}%**, "
        "significantly above the company average. Immediate intervention through targeted "
        "retention programs, manager training, and role satisfaction surveys is recommended "
        "for this department.\n\n"
        f"Compensation remains a critical driver of attrition. Employees in the lowest salary "
        f"band (<30K) leave at a rate of **{low_sal_rate:.1f}%** — nearly **{sal_multiple:.1f}x** "
        f"higher than the highest earners. Additionally, employees working overtime leave at "
        f"**{ot_attr:.1f}%** versus **{no_ot_attr:.1f}%** for those who don't, indicating "
        "workload imbalance. A salary review for low-band employees and overtime policy reform "
        "should be prioritized in the next budget cycle.\n\n"
        f"Tenure analysis reveals that employees in the **{high_ten}** bracket have the highest "
        f"attrition at **{ten_g['rate'].max():.1f}%**, suggesting a critical retention window. "
        f"The **{high_age}** age group also shows elevated attrition, pointing to career growth "
        "concerns among early-career employees. Implementing structured career development plans, "
        "mentorship programs, and promotion timelines for employees in their first two years "
        "could significantly reduce overall attrition."
    )


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────
def fig_to_png_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.getvalue()


def build_excel(df: pd.DataFrame, kpis: dict, insights: str, chart_pngs: dict) -> bytes:
    """Build a polished Excel workbook (Dashboard + KPIs + Raw + Summary)."""
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)

        summary_df = pd.DataFrame([
            ["Total Employees", kpis["total"]],
            ["Employees Left", kpis["attr_count"]],
            ["Attrition Rate", f"{kpis['attr_rate']:.1f}%"],
            ["Active Employees", kpis["total"] - kpis["attr_count"]],
            ["Avg Monthly Income", f"{df['MonthlyIncome'].mean():,.0f}"],
            ["Avg Age", f"{df['Age'].mean():.1f} yrs"],
            ["Avg Tenure", f"{df['YearsAtCompany'].mean():.1f} yrs"],
            ["OT Attrition Rate", f"{kpis['ot_attr']:.1f}%"],
            ["Non-OT Attrition Rate", f"{kpis['no_ot_attr']:.1f}%"],
            ["Highest Risk Department", kpis["high_dept"]],
            ["Highest Risk Age Group", kpis["high_age"]],
            ["Highest Risk Tenure", kpis["high_ten"]],
        ], columns=["Metric", "Value"])
        summary_df.to_excel(writer, sheet_name="KPI Summary", index=False)

        dept_full = df.groupby("Department").agg(
            Total=("AttritionNum", "count"),
            Left=("AttritionNum", "sum"),
            Avg_Salary=("MonthlyIncome", "mean"),
            Avg_Tenure=("YearsAtCompany", "mean"),
        ).round(1).reset_index()
        dept_full["Attrition_Rate"] = (dept_full["Left"] / dept_full["Total"] * 100).round(1)
        dept_full.to_excel(writer, sheet_name="Dept Analysis", index=False)

        pd.DataFrame({"Executive Summary": [insights]}).to_excel(
            writer, sheet_name="Executive Summary", index=False
        )

    buf.seek(0)
    wb = load_workbook(buf)

    # Dashboard sheet with charts embedded
    ws = wb.create_sheet("Dashboard", 0)
    DARK_BLUE = PatternFill("solid", fgColor="1B4F8A")
    ORANGE = PatternFill("solid", fgColor="D4380D")
    title_font = Font(color="FFFFFF", bold=True, name="Arial", size=16)

    for col, w in [(1, 3), (2, 22), (3, 18), (4, 18), (5, 18), (6, 18), (7, 3)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("B1:F2")
    ws["B1"] = "  HR ATTRITION ANALYSIS DASHBOARD"
    ws["B1"].fill = DARK_BLUE
    ws["B1"].font = title_font
    ws["B1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    kpi_cells = [
        ("TOTAL EMPLOYEES", kpis["total"], DARK_BLUE),
        ("ATTRITION RATE", f"{kpis['attr_rate']:.1f}%", ORANGE),
        ("EMPLOYEES LEFT", kpis["attr_count"], PatternFill("solid", fgColor="531DAB")),
        ("AVG TENURE", f"{df['YearsAtCompany'].mean():.1f} yrs",
         PatternFill("solid", fgColor="389E0D")),
    ]
    for col, (label, value, fill) in zip([2, 3, 4, 5], kpi_cells):
        cl = get_column_letter(col)
        ws[f"{cl}4"] = label
        ws[f"{cl}4"].fill = fill
        ws[f"{cl}4"].font = Font(color="FFFFFF", bold=True, name="Arial", size=8)
        ws[f"{cl}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{cl}5"] = value
        ws[f"{cl}5"].fill = fill
        ws[f"{cl}5"].font = Font(color="FFFFFF", bold=True, name="Arial", size=14)
        ws[f"{cl}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{cl}6"].fill = fill

    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[6].height = 14

    ws.merge_cells("B8:F8")
    ws["B8"] = "  ATTRITION ANALYSIS — 4 DIMENSIONS"
    ws["B8"].fill = PatternFill("solid", fgColor="2E4057")
    ws["B8"].font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    ws["B8"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[8].height = 18

    for png_bytes, anchor in [
        (chart_pngs["dept"], "B9"),
        (chart_pngs["age"], "B29"),
        (chart_pngs["salary"], "B49"),
        (chart_pngs["tenure"], "B69"),
    ]:
        img = XLImage(io.BytesIO(png_bytes))
        img.width, img.height = 430, 190
        ws.add_image(img, anchor)

    ins_row = 90
    ws.merge_cells(f"B{ins_row}:F{ins_row}")
    ws[f"B{ins_row}"] = "  EXECUTIVE SUMMARY"
    ws[f"B{ins_row}"].fill = DARK_BLUE
    ws[f"B{ins_row}"].font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    ws[f"B{ins_row}"].alignment = Alignment(vertical="center")
    ws.row_dimensions[ins_row].height = 18

    plain_insights = insights.replace("**", "")
    for j, para in enumerate(plain_insights.split("\n\n")[:3], 1):
        r = ins_row + j
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"] = para.strip()
        ws[f"B{r}"].font = Font(name="Arial", size=9, color="333333")
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("⚙️ Controls")
    seed = st.number_input("Random seed", value=42, min_value=0, max_value=9999, step=1,
                           help="Change the seed to regenerate the synthetic dataset.")
    n_emp = st.slider("Number of employees", 500, 3000, 1470, step=100)

    df_full = generate_dataset(n=n_emp, seed=int(seed))

    st.markdown("---")
    st.markdown("**Filters**")
    dept_options = sorted(df_full["Department"].unique())
    selected_depts = st.multiselect("Department", dept_options, default=dept_options)
    overtime_filter = st.radio("Overtime", ["All", "Yes only", "No only"], index=0)

    st.markdown("---")
    st.caption(
        "**About** — Synthetic IBM-style HR dataset. Attrition is generated from a "
        "rules-based probability model so trends are realistic and reproducible."
    )

# Apply filters
df = df_full[df_full["Department"].isin(selected_depts)].copy()
if overtime_filter == "Yes only":
    df = df[df["OverTime"] == "Yes"]
elif overtime_filter == "No only":
    df = df[df["OverTime"] == "No"]

if df.empty:
    st.error("No employees match the current filters. Widen the filters in the sidebar.")
    st.stop()

# Header
st.title("👥 HR Attrition Analytics")
st.markdown(
    "End-to-end people-analytics dashboard: generates a realistic 1,470-employee dataset, "
    "cleans it, identifies attrition drivers across 4 dimensions, and exports a "
    "boardroom-ready Excel report."
)

# KPIs
total = len(df)
attr_count = int(df["AttritionNum"].sum())
attr_rate = attr_count / total * 100
ot_attr = df[df["OverTime"] == "Yes"]["AttritionNum"].mean() * 100 if (df["OverTime"] == "Yes").any() else 0.0
no_ot_attr = df[df["OverTime"] == "No"]["AttritionNum"].mean() * 100 if (df["OverTime"] == "No").any() else 0.0

m1, m2, m3, m4 = st.columns(4)
m1.metric("Total Employees", f"{total:,}")
m2.metric("Attrition Rate", f"{attr_rate:.1f}%")
m3.metric("Employees Left", f"{attr_count:,}")
m4.metric("Avg Tenure", f"{df['YearsAtCompany'].mean():.1f} yrs")

st.markdown("---")

# Charts
fig1, dept_g = chart_dept(df, attr_rate)
fig2, age_g = chart_age(df, attr_rate)
fig3, sal_g = chart_salary(df)
fig4, ten_g = chart_tenure(df)

c1, c2 = st.columns(2)
with c1:
    st.pyplot(fig1, use_container_width=True)
    st.pyplot(fig3, use_container_width=True)
with c2:
    st.pyplot(fig2, use_container_width=True)
    st.pyplot(fig4, use_container_width=True)

st.markdown("---")

# Executive summary
high_dept = dept_g["rate"].idxmax()
high_age = age_g["rate"].idxmax()
high_ten = ten_g["rate"].idxmax()

insights = build_summary(df, dept_g, age_g, sal_g, ten_g)

st.subheader("📋 Executive Summary")
st.markdown(insights)

# Raw data + downloads
st.markdown("---")
st.subheader("📂 Data + Exports")

with st.expander("Show raw data table"):
    st.dataframe(df, use_container_width=True, hide_index=True)

kpis = {
    "total": total,
    "attr_count": attr_count,
    "attr_rate": attr_rate,
    "ot_attr": ot_attr,
    "no_ot_attr": no_ot_attr,
    "high_dept": high_dept,
    "high_age": high_age,
    "high_ten": high_ten,
}
chart_pngs = {
    "dept": fig_to_png_bytes(fig1),
    "age": fig_to_png_bytes(fig2),
    "salary": fig_to_png_bytes(fig3),
    "tenure": fig_to_png_bytes(fig4),
}
excel_bytes = build_excel(df, kpis, insights, chart_pngs)
csv_bytes = df.to_csv(index=False).encode("utf-8")

d1, d2 = st.columns(2)
d1.download_button(
    "⬇️  Download Excel report",
    data=excel_bytes,
    file_name="HR_Attrition_Analysis.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
d2.download_button(
    "⬇️  Download CSV",
    data=csv_bytes,
    file_name="hr_attrition_data.csv",
    mime="text/csv",
    use_container_width=True,
)
